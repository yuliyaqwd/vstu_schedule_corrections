from django.shortcuts import render
from django.http import HttpResponse
from django.views import View
from django.middleware.csrf import get_token
from .models import Correction, Item, ContextElement
import os
import tempfile
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from typing import List
import html

# Глобальная переменная: путь к последнему загруженному .xlsx файлу
last_uploaded_file_path = None


def _highlight_spaces(text: str) -> str:
    """Заменяет пробелы на полупрозрачные кружки · с подсветкой."""
    escaped = html.escape(text)
    return escaped.replace(' ', '<span class="space">·</span>')


def home(request):
    html_content = """
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Система корректировки расписаний</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 40px;
                line-height: 1.6;
                background: #f5f5f5;
            }
            .container {
                max-width: 1200px;
                margin: 0 auto;
                background: white;
                padding: 30px;
                border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }
            .header {
                background: #2c3e50;
                color: white;
                padding: 20px;
                border-radius: 5px;
                margin-bottom: 30px;
            }
            .menu {
                margin: 20px 0;
            }
            .menu a {
                display: inline-block;
                margin-right: 15px;
                padding: 10px 20px;
                background: #3498db;
                color: white;
                text-decoration: none;
                border-radius: 3px;
            }
            .menu a:hover {
                background: #2980b9;
            }
            .export-btn {
                background: #17a2b8 !important;
            }
            .schedule-export-btn {
                background: #6f42c1 !important;
            }
            .warning {
                background: #fff3cd;
                padding: 15px;
                border-radius: 5px;
                margin: 15px 0;
                border: 1px solid #ffeaa7;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Система корректировки учебных расписаний</h1>
                <p>Автоматизированная система для исправления ошибок в расписаниях</p>
            </div>

            <div class="menu">
                <a href="/corrections/">Таблица корректировок</a>
                <a href="/upload/">Загрузка расписания</a>
                <a href="/admin/">Админ-панель</a>
                <a href="/export/corrections/" class="export-btn">Экспорт корректировок</a>
                <a href="/export/schedule/" class="schedule-export-btn">Экспорт расписания</a>
            </div>

            <div class="info">
                <h2>Возможности системы:</h2>
                <ul>
                    <li>Автоматическое обнаружение ошибок в расписаниях</li>
                    <li>Ручное управление правилами корректировки</li>
                    <li>Визуализация изменений и истории</li>
                    <li>Загрузка Excel файлов в формате <strong>.xlsx</strong></li>
                    <li>Экспорт исправленных данных в Excel <strong>с сохранением всех стилей</strong></li>
                </ul>
                <p><strong>Важно:</strong> Поддерживается только формат <code>.xlsx</code> (Excel 2007+).</p>
            </div>
        </div>
    </body>
    </html>
    """
    return HttpResponse(html_content, content_type="text/html; charset=utf-8")


class CorrectionListView(View):
    def get(self, request):
        corrections = Correction.objects.all().order_by('-updated_at')
        table_rows = ""
        if corrections.exists():
            for correction in corrections:
                status_display = correction.get_status_display()
                status_color = self.get_status_color(correction.status)
                subject_value = _highlight_spaces(correction.subject.value)
                hypotheses = ', '.join([_highlight_spaces(h.value) for h in correction.hypotheses.all()])
                if not hypotheses:
                    hypotheses = "—"
                table_rows += f"""
                <tr style="{status_color}">
                    <td>{correction.id}</td>
                    <td>
                        <div class="monospace-cell" title="{html.escape(correction.subject.value)}">
                            {subject_value}
                        </div>
                    </td>
                    <td>{status_display}</td>
                    <td>
                        <div class="monospace-cell" title="{', '.join([h.value for h in correction.hypotheses.all()]) or '—'}">
                            {hypotheses}
                        </div>
                    </td>
                    <td>{correction.scope_id}</td>
                    <td>
                        <a href="/admin/corrections/correction/{correction.id}/change/" 
                           style="color: #3498db; text-decoration: none;">Редактировать</a>
                    </td>
                </tr>
                """
        else:
            table_rows = """
            <tr>
                <td colspan="6" style="text-align: center; padding: 40px; color: #666;">
                    Нет корректировок. Создайте их через админ-панель.
                </td>
            </tr>
            """
        has_uploaded_schedule = last_uploaded_file_path and os.path.exists(last_uploaded_file_path)
        html_content = f"""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Таблица корректировок</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 40px;
                    line-height: 1.6;
                    background: #f5f5f5;
                }}
                .container {{
                    max-width: 1400px;
                    margin: 0 auto;
                    background: white;
                    padding: 30px;
                    border-radius: 8px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                .header {{
                    background: #2c3e50;
                    color: white;
                    padding: 20px;
                    border-radius: 5px;
                    margin-bottom: 30px;
                }}
                .menu {{
                    margin: 20px 0;
                }}
                .menu a {{
                    display: inline-block;
                    margin-right: 15px;
                    padding: 10px 20px;
                    background: #3498db;
                    color: white;
                    text-decoration: none;
                    border-radius: 3px;
                }}
                .menu a:hover {{
                    background: #2980b9;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                    font-family: monospace;
                }}
                th, td {{
                    padding: 12px;
                    text-align: center;
                    border-bottom: 1px solid #ddd;
                    font-family: monospace !important;
                    vertical-align: top;
                }}
                th {{
                    background: #f8f9fa;
                    font-weight: bold;
                    font-family: monospace !important;
                    text-align: center;
                }}
                tr:hover {{
                    background: #f8f9fa;
                }}
                .monospace-cell {{
                    font-family: monospace !important;
                    text-align: center;
                    padding: 4px 6px;
                    border: 1px solid #eee;
                    border-radius: 3px;
                    background: transparent;
                    line-height: 1.4;
                    overflow: hidden;
                    text-overflow: ellipsis;
                    white-space: nowrap;
                    word-break: keep-all;
                }}
                .space {{
                    color: #aaa;
                    opacity: 0.7;
                    font-weight: bold;
                    background: rgba(255, 255, 255, 0.6);
                    padding: 0 1px;
                    border-radius: 2px;
                    font-size: 0.9em;
                }}
                .create-button {{
                    background: #28a745;
                    color: white;
                    padding: 10px 20px;
                    text-decoration: none;
                    border-radius: 3px;
                    display: inline-block;
                }}
                .create-button:hover {{
                    background: #218838;
                }}
                .export-btn {{
                    background: #17a2b8 !important;
                }}
                .schedule-export-btn {{
                    background: #6f42c1 !important;
                }}
                .warning {{
                    background: #fff3cd;
                    padding: 15px;
                    border-radius: 5px;
                    margin: 15px 0;
                    border: 1px solid #ffeaa7;
                }}
                .success {{
                    background: #d4edda;
                    padding: 15px;
                    border-radius: 5px;
                    margin: 15px 0;
                    border: 1px solid #c3e6cb;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>Таблица корректировок</h1>
                    <p>Все текстовые поля отображаются в моноширинном шрифте. Пробелы отмечены как <span style="font-family: monospace; color: #aaa;">·</span></p>
                </div>

                <div class="menu">
                    <a href="/">На главную</a>
                    <a href="/admin/corrections/correction/add/" class="create-button">Создать корректировку</a>
                    <a href="/upload/">Загрузить расписание</a>
                    <a href="/export/corrections/" class="export-btn">Экспорт корректировок</a>
                    <a href="/export/schedule/" class="schedule-export-btn">Экспорт расписания</a>
                </div>

                {f"<div class='success'><strong>Готово к экспорту:</strong> Расписание загружено (стили сохранены)</div>" if has_uploaded_schedule else "<div class='warning'><strong>Внимание:</strong> Загрузите файл в формате <code>.xlsx</code>, чтобы экспортировать расписание с корректировками и стилями.</div>"}

                <table>
                    <thead>
                        <tr>
                            <th>ID</th>
                            <th>Исходный предмет</th>
                            <th>Статус</th>
                            <th>Предлагаемые исправления</th>
                            <th>Scope</th>
                            <th>Действия</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows}
                    </tbody>
                </table>
            </div>
        </body>
        </html>
        """
        return HttpResponse(html_content, content_type="text/html; charset=utf-8")

    def get_status_color(self, status):
        if status == Correction.STATUS_PENDING:
            return "background-color: #fff9c4; color: #000;"
        elif status == Correction.STATUS_APPROVED:
            return "background-color: #e8f5e9; color: #000;"
        elif status == Correction.STATUS_INVALID:
            return "background-color: #ffcdd2; color: #000;"
        return ""


def upload_schedule(request):
    global last_uploaded_file_path

    if request.method == 'POST' and request.FILES.get('schedule_file'):
        uploaded_file = request.FILES['schedule_file']
        filename = uploaded_file.name.lower()

        if not filename.endswith('.xlsx'):
            return HttpResponse(
                "❌ Ошибка: поддерживается только формат <strong>.xlsx</strong> (Excel 2007 и новее).<br>"
                "Сохраните ваш файл как «Книга Excel (.xlsx)» в Microsoft Excel или LibreOffice и загрузите заново.",
                content_type="text/html; charset=utf-8"
            )

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                for chunk in uploaded_file.chunks():
                    tmp.write(chunk)
                last_uploaded_file_path = tmp.name

            wb_preview = load_workbook(last_uploaded_file_path, read_only=True, data_only=True)
            ws_preview = wb_preview.active
            rows = list(ws_preview.iter_rows(values_only=True))
            row_count = len(rows)
            col_count = max(len(row) for row in rows) if rows else 0
            wb_preview.close()

            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Результат загрузки</title>
                <style>
                    body {{ font-family: Arial; margin: 40px; background: #f5f5f5; }}
                    .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; }}
                    .header {{ background: #2c3e50; color: white; padding: 20px; border-radius: 5px; margin-bottom: 30px; }}
                    .success {{ background: #d4edda; padding: 15px; border-radius: 5px; margin: 15px 0; }}
                    .info-box {{ background: #e8f4fd; padding: 15px; border-radius: 5px; margin: 15px 0; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>✅ Файл успешно загружен!</h1>
                    </div>
                    <div class="success">
                        <p><strong>Имя файла:</strong> {uploaded_file.name}</p>
                        <p><strong>Формат:</strong> .xlsx</p>
                        <p><strong>Строк:</strong> {row_count}</p>
                        <p><strong>Колонок:</strong> {col_count}</p>
                        <p style="color: #155724;"><strong>✅ Все стили (цвета, шрифты, границы) сохранены!</strong></p>
                    </div>
                    <div style="margin-top: 20px;">
                        <a href="/upload/" style="background: #3498db; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px;">
                            Загрузить другой файл
                        </a>
                        <a href="/corrections/" style="background: #28a745; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; margin-left: 10px;">
                            Перейти к корректировкам
                        </a>
                        <a href="/export/schedule/" style="background: #6f42c1; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; margin-left: 10px;">
                            Экспорт исправленного расписания
                        </a>
                    </div>
                </div>
            </body>
            </html>
            """
            return HttpResponse(html_content, content_type="text/html; charset=utf-8")

        except Exception as e:
            if last_uploaded_file_path and os.path.exists(last_uploaded_file_path):
                os.unlink(last_uploaded_file_path)
            last_uploaded_file_path = None
            return HttpResponse(f"Ошибка при обработке файла: {str(e)}")

    csrf_token = get_token(request)
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Загрузка расписания</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 40px;
                line-height: 1.6;
                background: #f5f5f5;
            }}
            .container {{
                max-width: 800px;
                margin: 0 auto;
                background: white;
                padding: 30px;
                border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }}
            .header {{
                background: #2c3e50;
                color: white;
                padding: 20px;
                border-radius: 5px;
                margin-bottom: 30px;
            }}
            .upload-form {{
                background: #e8f4fd;
                padding: 25px;
                border-radius: 8px;
                border: 2px dashed #3498db;
                text-align: center;
            }}
            .file-input {{
                margin: 20px 0;
            }}
            .submit-btn {{
                background: #3498db;
                color: white;
                padding: 12px 30px;
                border: none;
                border-radius: 5px;
                font-size: 16px;
                cursor: pointer;
            }}
            .submit-btn:hover {{
                background: #2980b9;
            }}
            .info {{
                background: #fff3cd;
                padding: 15px;
                border-radius: 5px;
                margin-top: 20px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Загрузка файла расписания</h1>
                <p>Только <strong>.xlsx</strong> (Excel 2007+)</p>
            </div>
            <div class="upload-form">
                <form method="post" enctype="multipart/form-data">
                    <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                    <h3>Выберите файл Excel (.xlsx)</h3>
                    <div class="file-input">
                        <input type="file" name="schedule_file" accept=".xlsx" required style="font-size: 16px; padding: 10px;">
                    </div>
                    <button type="submit" class="submit-btn">Загрузить</button>
                </form>
            </div>
            <div class="info">
                <h4>Важно:</h4>
                <p>Поддерживается <strong>только формат .xlsx</strong>.</p>
                <p>При экспорте будут сохранены все стили: цвета, шрифты, границы и объединённые ячейки.</p>
            </div>
            <div style="text-align: center; margin-top: 20px;">
                <a href="/" style="color: #3498db; text-decoration: none;">← На главную</a>
            </div>
        </div>
    </body>
    </html>
    """
    return HttpResponse(html_content, content_type="text/html; charset=utf-8")


def apply_correction(subject: Item, hypotheses: List[Item], scope_id: int = 0) -> Item:
    from django.db.models import Q

    corrections = Correction.objects.filter(
        subject__value=subject.value,
        scope_id=scope_id
    ).select_related('subject').prefetch_related('hypotheses', 'subject__context')

    if not corrections.exists():
        if hypotheses:
            return max(hypotheses, key=lambda h: h.score)
        return subject

    correction = corrections.first()

    if correction.status == Correction.STATUS_APPROVED:
        approved_hyp = correction.hypotheses.filter(approved=True).first()
        current_score = approved_hyp.score if approved_hyp else -float('inf')
        new_better_hypotheses = [h for h in hypotheses if h.score > current_score]
        if new_better_hypotheses:
            for hyp in new_better_hypotheses:
                correction.hypotheses.add(hyp)
            correction.status = Correction.STATUS_PENDING
            correction.save()
        if approved_hyp:
            return approved_hyp
        best_in_correction = correction.hypotheses.order_by('-score').first()
        return best_in_correction or subject

    elif correction.status == Correction.STATUS_PENDING:
        for hyp in hypotheses:
            correction.hypotheses.add(hyp)
        correction.save()
        best = correction.hypotheses.order_by('-score').first()
        return best or subject

    elif correction.status == Correction.STATUS_INVALID:
        for hyp in correction.hypotheses.filter(suggested_by_reviewer=True):
            hyp.score = 0
            hyp.approved = False
            hyp.save()
        for hyp in hypotheses:
            correction.hypotheses.add(hyp)
        correction.status = Correction.STATUS_PENDING
        correction.save()
        if hypotheses:
            return max(hypotheses, key=lambda h: h.score)
        return subject

    return subject


def get_approved_correction_for_subject(subject_value: str, scope_id: int = 0) -> str:
    try:
        correction = Correction.objects.filter(
            subject__value=subject_value,
            scope_id=scope_id,
            status=Correction.STATUS_APPROVED
        ).select_related('subject').prefetch_related('hypotheses').first()

        if correction:
            approved_hyp = correction.hypotheses.filter(approved=True).first()
            if approved_hyp:
                return approved_hyp.value
    except Exception:
        pass
    return subject_value


def export_schedule_with_corrections(request):
    global last_uploaded_file_path

    if not last_uploaded_file_path or not os.path.exists(last_uploaded_file_path):
        html_content = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Ошибка экспорта</title>
            <style>
                body { font-family: Arial; margin: 40px; background: #f5f5f5; }
                .container { max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; }
                .header { background: #dc3545; color: white; padding: 20px; border-radius: 5px; margin-bottom: 30px; }
                .error { background: #f8d7da; padding: 15px; border-radius: 5px; margin: 15px 0; }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>❌ Ошибка экспорта</h1>
                </div>
                <div class="error">
                    <h3>Расписание не загружено</h3>
                    <p>Сначала загрузите файл в формате <strong>.xlsx</strong>.</p>
                </div>
                <div style="margin-top: 20px;">
                    <a href="/upload/" style="background: #3498db; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px;">
                        Перейти к загрузке
                    </a>
                    <a href="/" style="background: #6c757d; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; margin-left: 10px;">
                        На главную
                    </a>
                </div>
            </div>
        </body>
        </html>
        """
        return HttpResponse(html_content, content_type="text/html; charset=utf-8")

    output = BytesIO()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'schedule_corrected_{timestamp}.xlsx'

    try:
        wb = load_workbook(last_uploaded_file_path)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        original_value = str(cell.value).strip()
                        corrected_value = get_approved_correction_for_subject(original_value, scope_id=0)
                        cell.value = corrected_value

        wb.save(output)

    finally:
        if last_uploaded_file_path and os.path.exists(last_uploaded_file_path):
            os.unlink(last_uploaded_file_path)
        last_uploaded_file_path = None

    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_corrections(request):
    import pandas as pd
    corrections = Correction.objects.all().select_related('subject').prefetch_related('hypotheses', 'subject__context')
    data = []
    for correction in corrections:
        hypotheses = ', '.join([h.value for h in correction.hypotheses.all()])
        if not hypotheses:
            hypotheses = "—"
        context = ', '.join([f"{ctx.key}:{ctx.value}" for ctx in correction.subject.context.all()])
        if not context:
            context = "—"
        data.append({
            'ID': correction.id,
            'Исходный_предмет': correction.subject.value,
            'Статус': correction.get_status_display(),
            'Гипотезы': hypotheses,
            'Контекст': context,
            'Scope': correction.scope_id,
            'Создано': correction.created_at.strftime('%Y-%m-%d %H:%M'),
            'Обновлено': correction.updated_at.strftime('%Y-%m-%d %H:%M')
        })
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Корректировки', index=False)
        worksheet = writer.sheets['Корректировки']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    output.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'corrections_export_{timestamp}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
